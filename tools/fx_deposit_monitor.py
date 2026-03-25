"""
外币存款收益监控表生成器
======================================================
结构：客户美元 → 境外掉期(买外币/卖USD) → 境外分行外币存款
    → 到期掉期(买USD/卖外币) → 锁汇综合USD收益率

使用方法：
    python tools/fx_deposit_monitor.py
    或传入日期参数：
    python tools/fx_deposit_monitor.py --date 2026-03-03

输出：tools/fx_deposit_monitor_YYYYMMDD.xlsx
======================================================
"""

import argparse
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
except ImportError:
    print("请安装 openpyxl：pip install openpyxl")
    raise

# ─────────────────────────────────────────────
# 配置区
# ─────────────────────────────────────────────

TENORS = ["1M", "3M", "6M", "1Y"]
TENOR_DAYS = {"1M": 30, "3M": 91, "6M": 182, "1Y": 365}

# 监控币种 (代码, 全称, 报价单位说明)
CURRENCIES = [
    ("EUR", "欧元",       "USD/EUR，1EUR=?USD  (直接报价)"),
    ("GBP", "英镑",       "USD/GBP，1GBP=?USD  (直接报价)"),
    ("AUD", "澳元",       "USD/AUD，1AUD=?USD  (直接报价)"),
    ("CAD", "加元",       "CAD/USD，1USD=?CAD  (间接报价)"),
    ("JPY", "日元",       "JPY/USD，1USD=?JPY  (间接报价)"),
    ("HKD", "港元",       "HKD/USD，1USD=?HKD  (间接报价)"),
    ("SGD", "新加坡元",   "SGD/USD，1USD=?SGD  (间接报价)"),
    ("CHF", "瑞郎",       "USD/CHF，1CHF≈?USD  (直接报价)"),
    ("NZD", "纽元",       "USD/NZD，1NZD=?USD  (直接报价)"),
]

# 示例占位数据（每周手工更新黄色区域，此处仅为模板演示）
# 格式：{ccy: {tenor: {"spot": float, "dep_rate": float, "swap_pts_ann": float, "sofr_bench": float}}}
SAMPLE_DATA = {
    "EUR": {
        "1M":  {"spot": 1.0850, "dep_rate": 2.80, "swap_pts_ann": -1.85, "sofr_bench": 4.33},
        "3M":  {"spot": 1.0850, "dep_rate": 2.85, "swap_pts_ann": -1.82, "sofr_bench": 4.33},
        "6M":  {"spot": 1.0850, "dep_rate": 2.90, "swap_pts_ann": -1.79, "sofr_bench": 4.33},
        "1Y":  {"spot": 1.0850, "dep_rate": 2.95, "swap_pts_ann": -1.75, "sofr_bench": 4.33},
    },
    "GBP": {
        "1M":  {"spot": 1.2650, "dep_rate": 4.50, "swap_pts_ann": -0.22, "sofr_bench": 4.33},
        "3M":  {"spot": 1.2650, "dep_rate": 4.55, "swap_pts_ann": -0.20, "sofr_bench": 4.33},
        "6M":  {"spot": 1.2650, "dep_rate": 4.58, "swap_pts_ann": -0.18, "sofr_bench": 4.33},
        "1Y":  {"spot": 1.2650, "dep_rate": 4.60, "swap_pts_ann": -0.15, "sofr_bench": 4.33},
    },
    "AUD": {
        "1M":  {"spot": 0.6350, "dep_rate": 4.10, "swap_pts_ann":  0.15, "sofr_bench": 4.33},
        "3M":  {"spot": 0.6350, "dep_rate": 4.15, "swap_pts_ann":  0.12, "sofr_bench": 4.33},
        "6M":  {"spot": 0.6350, "dep_rate": 4.20, "swap_pts_ann":  0.10, "sofr_bench": 4.33},
        "1Y":  {"spot": 0.6350, "dep_rate": 4.25, "swap_pts_ann":  0.08, "sofr_bench": 4.33},
    },
    "CAD": {
        "1M":  {"spot": 1.4350, "dep_rate": 3.00, "swap_pts_ann": -0.05, "sofr_bench": 4.33},
        "3M":  {"spot": 1.4350, "dep_rate": 3.05, "swap_pts_ann": -0.04, "sofr_bench": 4.33},
        "6M":  {"spot": 1.4350, "dep_rate": 3.10, "swap_pts_ann": -0.03, "sofr_bench": 4.33},
        "1Y":  {"spot": 1.4350, "dep_rate": 3.15, "swap_pts_ann": -0.02, "sofr_bench": 4.33},
    },
    "JPY": {
        "1M":  {"spot": 150.50, "dep_rate": 0.10, "swap_pts_ann":  4.20, "sofr_bench": 4.33},
        "3M":  {"spot": 150.50, "dep_rate": 0.12, "swap_pts_ann":  4.10, "sofr_bench": 4.33},
        "6M":  {"spot": 150.50, "dep_rate": 0.15, "swap_pts_ann":  4.00, "sofr_bench": 4.33},
        "1Y":  {"spot": 150.50, "dep_rate": 0.18, "swap_pts_ann":  3.85, "sofr_bench": 4.33},
    },
    "HKD": {
        "1M":  {"spot": 7.7850, "dep_rate": 4.00, "swap_pts_ann":  0.30, "sofr_bench": 4.33},
        "3M":  {"spot": 7.7850, "dep_rate": 4.05, "swap_pts_ann":  0.28, "sofr_bench": 4.33},
        "6M":  {"spot": 7.7850, "dep_rate": 4.10, "swap_pts_ann":  0.25, "sofr_bench": 4.33},
        "1Y":  {"spot": 7.7850, "dep_rate": 4.15, "swap_pts_ann":  0.22, "sofr_bench": 4.33},
    },
    "SGD": {
        "1M":  {"spot": 1.3450, "dep_rate": 3.20, "swap_pts_ann":  0.90, "sofr_bench": 4.33},
        "3M":  {"spot": 1.3450, "dep_rate": 3.25, "swap_pts_ann":  0.88, "sofr_bench": 4.33},
        "6M":  {"spot": 1.3450, "dep_rate": 3.30, "swap_pts_ann":  0.85, "sofr_bench": 4.33},
        "1Y":  {"spot": 1.3450, "dep_rate": 3.35, "swap_pts_ann":  0.80, "sofr_bench": 4.33},
    },
    "CHF": {
        "1M":  {"spot": 0.9050, "dep_rate": 0.50, "swap_pts_ann":  3.75, "sofr_bench": 4.33},
        "3M":  {"spot": 0.9050, "dep_rate": 0.55, "swap_pts_ann":  3.70, "sofr_bench": 4.33},
        "6M":  {"spot": 0.9050, "dep_rate": 0.60, "swap_pts_ann":  3.65, "sofr_bench": 4.33},
        "1Y":  {"spot": 0.9050, "dep_rate": 0.65, "swap_pts_ann":  3.55, "sofr_bench": 4.33},
    },
    "NZD": {
        "1M":  {"spot": 0.5750, "dep_rate": 3.50, "swap_pts_ann":  0.72, "sofr_bench": 4.33},
        "3M":  {"spot": 0.5750, "dep_rate": 3.55, "swap_pts_ann":  0.70, "sofr_bench": 4.33},
        "6M":  {"spot": 0.5750, "dep_rate": 3.60, "swap_pts_ann":  0.68, "sofr_bench": 4.33},
        "1Y":  {"spot": 0.5750, "dep_rate": 3.65, "swap_pts_ann":  0.65, "sofr_bench": 4.33},
    },
}

# ─────────────────────────────────────────────
# 颜色/样式常量
# ─────────────────────────────────────────────
COLOR_NAVY      = "1F3864"   # 深蓝 标题背景
COLOR_DARK_BLUE = "2E5596"   # 蓝   副标题背景
COLOR_HEADER    = "4472C4"   # 蓝   列头背景
COLOR_SUB_HDR   = "BDD7EE"   # 浅蓝 子列头背景
COLOR_INPUT_BG  = "FFFACD"   # 淡黄 可编辑输入区
COLOR_CALC_BG   = "EBF3FB"   # 浅蓝 计算结果区
COLOR_GREEN     = "C6EFCE"   # 绿   优秀
COLOR_YELLOW_L  = "FFEB9C"   # 黄   一般
COLOR_RED_L     = "FFC7CE"   # 红   较差
COLOR_FONT_W    = "FFFFFF"
COLOR_FONT_D    = "1F3864"


def make_font(bold=False, size=10, color=COLOR_FONT_D, italic=False):
    return Font(name="微软雅黑", bold=bold, size=size, color=color, italic=italic)


def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def make_border(thin_all=False):
    side = Side(style="thin", color="B8CCE4")
    side_med = Side(style="medium", color="4472C4")
    if thin_all:
        return Border(left=side, right=side, top=side, bottom=side)
    return Border(left=side, right=side, top=side, bottom=side)


def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def right_align():
    return Alignment(horizontal="right", vertical="center")


# ─────────────────────────────────────────────
# 主生成函数
# ─────────────────────────────────────────────

def build_workbook(report_date: date) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── Sheet 1: 监控主表 ──────────────────────
    ws_main = wb.active
    ws_main.title = "监控主表"
    _build_main_sheet(ws_main, report_date)

    # ── Sheet 2: 说明与公式 ───────────────────
    ws_note = wb.create_sheet("说明与公式")
    _build_note_sheet(ws_note)

    return wb


# ─────────────────────────────────────────────
# Sheet 1: 监控主表
# ─────────────────────────────────────────────
def _build_main_sheet(ws, report_date: date):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B7"          # 冻结首列+前6行

    n_ccy = len(CURRENCIES)
    n_tenor = len(TENORS)

    # ── 列宽设置 ───────────────────────────────
    ws.column_dimensions["A"].width = 14   # 币种列
    for col_idx in range(2, 2 + n_tenor * 5 + 5):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11.5

    # ── 行高 ───────────────────────────────────
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 22
    for r in range(7, 7 + n_ccy + 2):
        ws.row_dimensions[r].height = 22

    # ── TITLE ─────────────────────────────────
    total_cols = 1 + n_tenor * 5 + 1   # A + 每tenor5列 + 备注列
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    title_cell = ws.cell(2, 1,
        value=f"外币存款综合收益监控表   |   更新日期：{report_date.strftime('%Y-%m-%d')} (周{['一','二','三','四','五','六','日'][report_date.weekday()]})")
    title_cell.font  = make_font(bold=True, size=16, color=COLOR_FONT_W)
    title_cell.fill  = make_fill(COLOR_NAVY)
    title_cell.alignment = center()

    # ── 副标题 ────────────────────────────────
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    sub = ws.cell(3, 1,
        value="结构：客户USD  →  境外掉期(卖USD买外币)  →  境外分行外币存款  →  到期掉期(买USD卖外币)  →  综合USD收益率锁定")
    sub.font = make_font(italic=True, size=10, color=COLOR_FONT_W)
    sub.fill = make_fill(COLOR_DARK_BLUE)
    sub.alignment = center()

    # ── 说明行 ────────────────────────────────
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)
    note = ws.cell(4, 1,
        value="【黄底单元格为手动录入区】  综合USD收益 = 外币存款利率(%) + 掉期换汇折年收益(%)    掉期：境外机构    存款：境外分行询价    基准：SOFR")
    note.font = make_font(size=9, color="7F3F00")
    note.fill = make_fill("FFF2CC")
    note.alignment = center(wrap=True)

    # ── 期限大标题行（第5行）────────────────
    ws.cell(5, 1, value="币种").font = make_font(bold=True, size=10, color=COLOR_FONT_W)
    ws.cell(5, 1).fill      = make_fill(COLOR_HEADER)
    ws.cell(5, 1).alignment = center()
    ws.cell(5, 1).border    = make_border()

    col = 2
    for tenor in TENORS:
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 4)
        hdr = ws.cell(5, col, value=f"◆  {tenor}  ◆")
        hdr.font      = make_font(bold=True, size=11, color=COLOR_FONT_W)
        hdr.fill      = make_fill(COLOR_HEADER)
        hdr.alignment = center()
        hdr.border    = make_border()
        col += 5

    # 备注列
    ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
    remark_hdr = ws.cell(5, col, value="备注 / 询价信息")
    remark_hdr.font      = make_font(bold=True, size=9, color=COLOR_FONT_W)
    remark_hdr.fill      = make_fill(COLOR_HEADER)
    remark_hdr.alignment = center(wrap=True)
    remark_hdr.border    = make_border()

    # ── 子列头行（第6行）────────────────────
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=1)
    loc = ws.cell(6, 1, value="")
    loc.fill = make_fill(COLOR_HEADER)

    sub_headers = [
        ("即期汇率\n(报价)",          COLOR_SUB_HDR, False),
        ("外币存款\n利率(%)",          COLOR_INPUT_BG, True),
        ("掉期折年\n收益(bp)",         COLOR_INPUT_BG, True),
        ("综合USD\n收益率(%)",         COLOR_CALC_BG, False),
        ("超额收益\nvs SOFR(bp)",      COLOR_CALC_BG, False),
    ]
    col = 2
    for tenor in TENORS:
        for (sh, bg, is_input) in sub_headers:
            c = ws.cell(6, col, value=sh)
            c.font      = make_font(bold=True, size=8.5,
                                    color=COLOR_FONT_D if not is_input else "7F3F00")
            c.fill      = make_fill(bg)
            c.alignment = center(wrap=True)
            c.border    = make_border()
            col += 1

    # ── 数据行（第7行起）────────────────────
    DATA_START_ROW = 7
    for r_idx, (ccy, ccy_name, _) in enumerate(CURRENCIES):
        row = DATA_START_ROW + r_idx
        d = SAMPLE_DATA.get(ccy, {})

        # 币种列
        ccy_cell = ws.cell(row, 1, value=f"{ccy}\n{ccy_name}")
        ccy_cell.font      = make_font(bold=True, size=10, color=COLOR_FONT_D)
        ccy_cell.fill      = make_fill("DCE6F1")
        ccy_cell.alignment = center(wrap=True)
        ccy_cell.border    = make_border()

        col = 2
        for tenor in TENORS:
            td = d.get(tenor, {})
            spot         = td.get("spot", 0.0)
            dep_rate     = td.get("dep_rate", 0.0)
            swap_pts_ann = td.get("swap_pts_ann", 0.0)  # 年化bp
            sofr_bench   = td.get("sofr_bench", 4.33)
            total_return = dep_rate + swap_pts_ann / 100.0   # 综合%
            excess_bp    = round((total_return - sofr_bench) * 100, 2)

            # 即期汇率（蓝底，手动录入，但初始有示例值）
            c_spot = ws.cell(row, col, value=spot)
            c_spot.font      = make_font(size=9)
            c_spot.fill      = make_fill(COLOR_INPUT_BG)
            c_spot.alignment = center()
            c_spot.border    = make_border()
            c_spot.number_format = "0.0000"
            col += 1

            # 外币存款利率（黄底手动）
            c_dep = ws.cell(row, col, value=dep_rate)
            c_dep.font      = make_font(size=9, color="7F3F00")
            c_dep.fill      = make_fill(COLOR_INPUT_BG)
            c_dep.alignment = center()
            c_dep.border    = make_border()
            c_dep.number_format = "0.00"
            col += 1

            # 掉期折年bp（黄底手动）
            c_swap = ws.cell(row, col, value=swap_pts_ann)
            c_swap.font      = make_font(size=9, color="7F3F00")
            c_swap.fill      = make_fill(COLOR_INPUT_BG)
            c_swap.alignment = center()
            c_swap.border    = make_border()
            c_swap.number_format = '0.00;[Red]-0.00'
            col += 1

            # 综合USD收益率（公式）
            dep_col  = get_column_letter(col - 2)
            swap_col = get_column_letter(col - 1)
            formula_total = f"={dep_col}{row}+{swap_col}{row}/100"
            c_total = ws.cell(row, col, value=total_return)
            c_total.font        = make_font(bold=True, size=9)
            c_total.fill        = make_fill(COLOR_CALC_BG)
            c_total.alignment   = center()
            c_total.border      = make_border()
            c_total.number_format = "0.00"
            col += 1

            # 超额收益 vs SOFR（bp）
            total_col  = get_column_letter(col - 1)
            formula_ex = f"=({total_col}{row}-{sofr_bench})*100"
            c_excess = ws.cell(row, col, value=excess_bp)
            c_excess.font      = make_font(bold=True, size=9)
            c_excess.fill      = make_fill(COLOR_CALC_BG)
            c_excess.alignment = center()
            c_excess.border    = make_border()
            c_excess.number_format = '+0.0;-0.0;0.0'
            col += 1

    # ── 备注列（数据区）────────────────────
    remark_col = 2 + n_tenor * 5
    for r_idx in range(n_ccy):
        row = DATA_START_ROW + r_idx
        rc = ws.cell(row, remark_col, value="—")
        rc.font      = make_font(size=9, color="595959")
        rc.fill      = make_fill("F2F2F2")
        rc.alignment = center(wrap=True)
        rc.border    = make_border()

    # ── SOFR基准行 ───────────────────────────
    bench_row = DATA_START_ROW + n_ccy + 1
    ws.row_dimensions[bench_row].height = 20
    ws.merge_cells(start_row=bench_row, start_column=1, end_row=bench_row, end_column=2)
    b0 = ws.cell(bench_row, 1, value="SOFR基准利率 (%)")
    b0.font      = make_font(bold=True, size=9, color=COLOR_FONT_W)
    b0.fill      = make_fill(COLOR_NAVY)
    b0.alignment = center()
    b0.border    = make_border()

    col = 3
    sofr_val = 4.33
    for tenor in TENORS:
        ws.merge_cells(start_row=bench_row, start_column=col, end_row=bench_row, end_column=col + 3)
        bc = ws.cell(bench_row, col, value=sofr_val)
        bc.font        = make_font(bold=True, size=10, color=COLOR_FONT_W)
        bc.fill        = make_fill(COLOR_NAVY)
        bc.alignment   = center()
        bc.border      = make_border()
        bc.number_format = "0.00"
        col += 5

    # ── 条件格式：超额收益列高亮 ─────────────
    # 每个tenor的第5列（超额收益）做颜色条件格式
    col_idx_check = 2 + 4   # 第1个tenor的第5列
    for t_idx in range(n_tenor):
        exc_col = get_column_letter(2 + t_idx * 5 + 4)
        cell_range = f"{exc_col}{DATA_START_ROW}:{exc_col}{DATA_START_ROW + n_ccy - 1}"
        # 正值（超额收益>0）→ 绿
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=make_fill(COLOR_GREEN),
                       font=Font(name="微软雅黑", bold=True, size=9, color="375623"))
        )
        # 负值 → 红
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=make_fill(COLOR_RED_L),
                       font=Font(name="微软雅黑", bold=True, size=9, color="9C0006"))
        )

    # ── Tab颜色 ───────────────────────────────
    ws.sheet_properties.tabColor = "1F3864"

    # ── 打印设置 ──────────────────────────────
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.paperSize     = 9       # A4
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 0
    ws.print_title_rows = "1:6"


# ─────────────────────────────────────────────
# Sheet 2: 说明与公式
# ─────────────────────────────────────────────
def _build_note_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "4472C4"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 30

    rows_data = [
        (None, None, ""),
        ("TITLE", None, "外币存款综合收益监控表 — 使用说明 & 公式逻辑"),
        (None, None, ""),
        # 核心公式
        ("HEAD", None, "一、核心计算公式"),
        ("SUB",  "综合USD收益率",
         "= 外币存款利率(%) + 掉期折年收益(%)"),
        ("SUB",  "掉期折年收益(bp→%)",
         "= 掉期年化bp ÷ 100\n（填入时已折年，即期×(远期-即期)/即期 × 365/天数 × 10000）"),
        ("SUB",  "超额收益 vs SOFR",
         "= (综合USD收益率 − SOFR基准) × 100   单位：bp"),
        (None, None, ""),
        # 操作说明
        ("HEAD", None, "二、每周更新操作（周一 10:00前）"),
        ("STEP", "Step 1",  "向境外分行发询价：各币种 1M/3M/6M/1Y 外币存款利率"),
        ("STEP", "Step 2",  "向境外掉期交易对手（多家）询价：各币种各期限掉期报价"),
        ("STEP", "Step 3",  "查彭博/路透获取当日即期汇率（9:30 Fix 或 实时Mid）"),
        ("STEP", "Step 4",  "填入监控主表黄底单元格（即期汇率 / 外币存款利率 / 掉期折年bp）"),
        ("STEP", "Step 5",  "绿色单元格自动计算综合收益率 & 超额收益，高亮显示最优币种"),
        ("STEP", "Step 6",  "截图发给肖旭，并附最优1-2个币种建议 + 境外分行存款报价单"),
        (None, None, ""),
        # 产品要点
        ("HEAD", None, "三、产品结构要点"),
        ("ITEM", "客户端",   "持有USD的理财产品客户（固定USD规模）"),
        ("ITEM", "资产端",   "境外分行外币存款（优选），CD流动性较低需多家凑单并持有至到期"),
        ("ITEM", "掉期端",   "境外机构交易商（已有授信），买卖掉期锁定汇率风险"),
        ("ITEM", "结算端",   "全程USD进出，外汇风险通过掉期完全对冲"),
        ("ITEM", "授信",     "金市境内暂无授信额度 → 走境外掉期，该通道已打通"),
        (None, None, ""),
        # 颜色说明
        ("HEAD", None, "四、颜色含义"),
        ("COLOR", "🟡 黄底单元格",  "手动录入区——每周更新即期汇率、存款利率、掉期bp"),
        ("COLOR", "🔵 蓝底单元格",  "自动计算区——综合收益率 & 超额收益（勿手动修改）"),
        ("COLOR", "🟢 绿色高亮",    "超额收益 > 0bp，优于SOFR，可推荐客户"),
        ("COLOR", "🔴 红色高亮",    "超额收益 < 0bp，低于SOFR，不建议"),
        (None, None, ""),
        ("NOTE", None,
         "免责：本表为内部工作监控工具，数据需实时询价核实，不构成对外报价。"),
    ]

    for r, row_data in enumerate(rows_data, start=1):
        ws.row_dimensions[r].height = 20
        style_type, label, content = row_data

        if style_type == "TITLE":
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            c = ws.cell(r, 2, value=content)
            c.font = make_font(bold=True, size=14, color=COLOR_FONT_W)
            c.fill = make_fill(COLOR_NAVY)
            c.alignment = center()
        elif style_type == "HEAD":
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            c = ws.cell(r, 2, value=content)
            c.font = make_font(bold=True, size=11, color=COLOR_FONT_W)
            c.fill = make_fill(COLOR_DARK_BLUE)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        elif style_type in ("SUB", "STEP", "ITEM", "COLOR"):
            c_lbl = ws.cell(r, 2, value=label)
            c_lbl.font = make_font(bold=True, size=9, color=COLOR_DARK_BLUE)
            c_lbl.fill = make_fill("DEEAF1")
            c_lbl.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            c_lbl.border = make_border()

            c_cnt = ws.cell(r, 3, value=content)
            c_cnt.font = make_font(size=9)
            c_cnt.fill = make_fill("F2F9FF")
            c_cnt.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True, indent=1)
            c_cnt.border = make_border()
            ws.row_dimensions[r].height = 28 if "\n" in content else 20
        elif style_type == "NOTE":
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            c = ws.cell(r, 2, value=content)
            c.font = make_font(italic=True, size=8.5, color="595959")
            c.fill = make_fill("F2F2F2")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


# ─────────────────────────────────────────────
# CLI 入口
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="生成外币存款收益监控Excel")
    parser.add_argument("--date", default=None,
                        help="报告日期，格式 YYYY-MM-DD，默认今天")
    parser.add_argument("--out",  default=None,
                        help="输出文件路径，默认 tools/fx_deposit_monitor_YYYYMMDD.xlsx")
    args = parser.parse_args()

    if args.date:
        report_date = datetime.strptime(args.date, "%Y-%m-%d").date()
    else:
        report_date = date.today()

    out_dir = Path(__file__).parent
    if args.out:
        out_path = Path(args.out)
    else:
        fname = f"fx_deposit_monitor_{report_date.strftime('%Y%m%d')}.xlsx"
        out_path = out_dir / fname

    wb = build_workbook(report_date)
    wb.save(str(out_path))
    print(f"✅ 已生成：{out_path.resolve()}")
    print(f"   📋 包含 Sheet：监控主表（黄底手录 + 绿底自算）、说明与公式")
    print(f"   📅 报告日期：{report_date.strftime('%Y-%m-%d')}")
    print(f"   🔄 每周一运行：python tools/fx_deposit_monitor.py --date YYYY-MM-DD")


if __name__ == "__main__":
    main()
