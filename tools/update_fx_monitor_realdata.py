"""
外币存款综合收益监控表 — 实时数据版 v2
================================================
全部数据源改用 FRED（联储圣路易斯）CSV API
  verify=False 绕过本机SSL证书问题
  不依赖 yfinance / ECB API / MAS API

即期汇率  FRED DEXUSEU / DEXJPUS 等
基准利率  FRED SOFR + 各币种interbank/政策利率
远期汇率  CIP 利率平价公式推算（理论值）
存款利率  = 公开基准参考；I列橙色留空 = 需境外分行询价
================================================
用法：
  python tools/update_fx_monitor_realdata.py
"""

import os, sys
import urllib3
urllib3.disable_warnings()
import requests
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.exit("请安装 openpyxl：pip install openpyxl")

OUT_DIR  = Path(r"c:\Users\pynbj\OneDrive\1.文档-积累要看的文件\1. 投资框架\11.投资机会跟踪报告")
OUT_FILE = OUT_DIR / f"外币存款收益监控表_{date.today().strftime('%Y%m%d')}.xlsx"
FRED_BASE = "https://fred.stlouisfed.org/graph/fredgraph.csv"
TENORS = [("3M", 90), ("6M", 180), ("9M", 270), ("12M", 360)]

CURRENCIES = [
    {"name":"EUR","pair":"EUR/USD","type":"FX_BASE","day_basis":360,
     "fred_spot":"DEXUSEU","fred_rate":"ECBDFR","rate_label":"ECB存款便利利率"},
    {"name":"GBP","pair":"GBP/USD","type":"FX_BASE","day_basis":365,
     "fred_spot":"DEXUSUK","fred_rate":"IUDSOIA","rate_label":"BOE SONIA"},
    {"name":"AUD","pair":"AUD/USD","type":"FX_BASE","day_basis":365,
     "fred_spot":"DEXUSAL","fred_rate":"IR3TIB01AUM156N","rate_label":"AUD 3M Interbank"},
    {"name":"NZD","pair":"NZD/USD","type":"FX_BASE","day_basis":365,
     "fred_spot":"DEXUSNZ","fred_rate":"IR3TIB01NZM156N","rate_label":"NZD 3M Interbank"},
    {"name":"CAD","pair":"USD/CAD","type":"USD_BASE","day_basis":360,
     "fred_spot":"DEXCAUS","fred_rate":"IR3TIB01CAM156N","rate_label":"CAD 3M Interbank"},
    {"name":"CHF","pair":"USD/CHF","type":"USD_BASE","day_basis":360,
     "fred_spot":"DEXSZUS","fred_rate":"IR3TIB01CHM156N","rate_label":"CHF 3M Interbank"},
    {"name":"JPY","pair":"USD/JPY","type":"USD_BASE","day_basis":360,
     "fred_spot":"DEXJPUS","fred_rate":"IRSTCB01JPM156N","rate_label":"BOJ政策利率"},
    {"name":"HKD","pair":"USD/HKD","type":"USD_BASE","day_basis":360,
     "fred_spot":"DEXHKUS","fred_rate":"IR3TIB01HKM156N","rate_label":"HIBOR 3M"},
    {"name":"SGD","pair":"USD/SGD","type":"USD_BASE","day_basis":360,
     "fred_spot":"DEXSIUS","fred_rate":"IR3TIB01SGM156N","rate_label":"SGD 3M Interbank"},
]

def fred_latest(series_id, divisor=1.0, label=""):
    url = f"{FRED_BASE}?id={series_id}"
    try:
        r = requests.get(url, verify=False, timeout=15)
        if not r.ok:
            print(f"  ERR {label or series_id}: HTTP {r.status_code}")
            return None, None
        lines = r.text.strip().split("\n")
        for line in reversed(lines[1:]):
            parts = line.strip().split(",")
            if len(parts)==2 and parts[1] not in (".","","NA"):
                try:
                    val = round(float(parts[1])/divisor, 6)
                    return val, parts[0]
                except:
                    pass
        print(f"  ERR {label or series_id}: no valid data")
        return None, None
    except Exception as e:
        print(f"  ERR {label or series_id}: {e}")
        return None, None

def fetch_all_data():
    print("="*58)
    print(f"抓取开始  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*58)
    results={}
    print("\n[1] SOFR")
    sofr,sofr_date=fred_latest("SOFR",divisor=100,label="SOFR")
    if sofr: print(f"  OK SOFR={sofr:.4%}  ({sofr_date})")
    results["SOFR"]={"value":sofr,"date":sofr_date}

    print("\n[2] 即期汇率 FRED")
    for ccy in CURRENCIES:
        name=ccy["name"]
        results[name]={}
        spot,spot_date=fred_latest(ccy["fred_spot"],label=f"{ccy['pair']} spot")
        dp=4 if name!="JPY" else 2
        if spot: print(f"  OK {ccy['pair']:9s}={spot:.{dp}f}  ({spot_date})")
        else:    print(f"  MISS {ccy['pair']:9s} 留空标橙")
        results[name]["spot"]=spot
        results[name]["spot_date"]=spot_date
        results[name]["spot_src"]=f"FRED/{ccy['fred_spot']}"

    print("\n[3] 基准利率 FRED")
    FALLBACK={
        "EUR":(0.0265,"ECB估算"),"GBP":(0.0473,"BOE估算"),
        "AUD":(0.0410,"RBA估算"),"NZD":(0.0350,"RBNZ估算"),
        "CAD":(0.0300,"BOC估算"),"CHF":(0.0050,"SNB估算"),
        "JPY":(0.0025,"BOJ估算"),"HKD":(0.0450,"HIBOR估算"),
        "SGD":(0.0325,"SORA估算"),
    }
    for ccy in CURRENCIES:
        name=ccy["name"]
        rate,rate_date=fred_latest(ccy["fred_rate"],divisor=100,label=f"{name} {ccy['rate_label']}")
        if rate:
            print(f"  OK {name:4s} {ccy['rate_label']:22s}={rate:.4%}  ({rate_date})")
        else:
            fv,fn=FALLBACK.get(name,(None,""))
            if fv:
                rate,rate_date=fv,fn
                print(f"  FB {name:4s} 估算={rate:.4%} ({fn})")
            else:
                print(f"  ERR {name:4s} 无数据")
        results[name]["rate"]=rate
        results[name]["rate_date"]=rate_date
        results[name]["rate_src"]=ccy["rate_label"]
    return results

def calc_fwd(spot,r_usd,r_fx,days,day_basis=360,ccy_type="FX_BASE"):
    if None in (spot,r_usd,r_fx) or spot==0: return None
    if ccy_type=="FX_BASE":
        return round(spot*(1+r_usd*days/360)/(1+r_fx*days/day_basis),6)
    else:
        return round(spot*(1+r_fx*days/day_basis)/(1+r_usd*days/360),6)

C_NAVY="1F3864";C_BLUE="2E5FA3";C_LT="D6E4F0";C_ODD="EBF2FA";C_EVEN="FFFFFF"
C_INPUT="FFFDE7";C_FORMULA="F0FFF4";C_CIP="E8F4FD";C_MISS="FFE0B2"

def F(sz=9,bold=False,color="1F1F1F",italic=False):
    return Font(name="微软雅黑",size=sz,bold=bold,color=color,italic=italic)
def A(h="center",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def Fill(c):
    return PatternFill("solid",fgColor=c)
def Bdr(bot="thin",bc="BFBFBF"):
    s=Side(style="thin",color="BFBFBF")
    sb=Side(style=bot,color=bc)
    return Border(top=s,bottom=sb,left=s,right=s)

def sc(ws,r,c,val=None,**kw):
    cell=ws.cell(r,c)
    if val is not None: cell.value=val
    if "font" in kw: cell.font=kw["font"]
    if "align" in kw: cell.alignment=kw["align"]
    if "bg" in kw: cell.fill=Fill(kw["bg"])
    if "border" in kw: cell.border=kw["border"]
    if "fmt" in kw: cell.number_format=kw["fmt"]
    return cell

def build_excel(data):
    wb=openpyxl.Workbook()
    ws=wb.active; ws.title="监控主表"
    ws.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFGHIJKLM",[5,7,9,6,6,12,13,11,13,13,11,14,22]):
        ws.column_dimensions[col].width=w

    today=date.today().strftime("%Y-%m-%d")
    now=datetime.now().strftime("%H:%M")
    sofr=data["SOFR"]["value"]; sofr_date=data["SOFR"]["date"]or"—"

    for r,h in [(1,6),(2,32),(3,16),(4,26),(5,32),(6,18)]:
        ws.row_dimensions[r].height=h

    ws.merge_cells("A2:M2")
    sc(ws,2,1,"外币存款综合收益监控表（境外掉期+境外分行存款）",
       font=F(14,bold=True,color=C_NAVY),align=A(),bg="EBF2FA")

    ws.merge_cells("A3:D3")
    sc(ws,3,1,f"更新：{today} {now}  |  汇率来源：FRED（T-1）",
       font=F(8,italic=True,color="595959"),align=A("left"))
    sc(ws,3,5,"SOFR基准(%)",font=F(8,bold=True),align=A("right"))
    sc(ws,3,6,sofr if sofr else 0.043,
       font=F(9,bold=True,color="C00000"),align=A(),bg=C_INPUT,fmt="0.00%")
    ws.merge_cells("G3:H3")
    sc(ws,3,7,f"FRED/SOFR ({sofr_date})",font=F(8,italic=True),align=A("left"))
    ws.merge_cells("I3:K3")
    sc(ws,3,9,"🟠橙色=无公开数据，向境外分行询价手动填入",
       font=F(8,bold=True,color="E65100"),align=A(),bg=C_MISS)
    ws.merge_cells("L3:M3")
    sc(ws,3,12,"I列=公开基准参考，需覆盖为实际询价",
       font=F(8,italic=True,color="595959"),align=A("left"))

    for rng,txt,clr in [("A4:E4","基本信息",C_NAVY),("F4:H4","掉期数据（境外询价）",C_BLUE),
                         ("I4:I4","境外存款",C_BLUE),("J4:K4","综合收益（自动）",C_NAVY),
                         ("L4:M4","备注",C_BLUE)]:
        ws.merge_cells(rng);c=ws[rng.split(":")[0]]
        c.value=txt;c.font=F(9,bold=True,color="FFFFFF");c.alignment=A();c.fill=Fill(clr)

    hdrs=["序","币种","报价对","期限","天数","即期\n(Spot\nFRED)","远期\n(Forward\nCIP理论)",
          "掉期年化\n(%)","存款报价\n(%)\n★填询价","综合USD\n收益(%)","较SOFR\n利差(bp)","数据来源","分行/备注"]
    for j,h in enumerate(hdrs,1):
        sc(ws,5,j,h,font=F(8,bold=True,color="FFFFFF"),align=A(wrap=True),bg=C_NAVY,border=Bdr())

    ws.merge_cells("A6:C6")
    sc(ws,6,1,"🟡黄-手填  🔵蓝-CIP参考  🟢绿-自动  🟠橙-询价",font=F(8,italic=True),align=A("left"))

    DATA_START=7; row=DATA_START; seq=1
    r_usd=sofr if sofr else 0.043
    n_tenor=len(TENORS)

    for ci,ccy in enumerate(CURRENCIES):
        bg=C_ODD if ci%2==0 else C_EVEN
        name=ccy["name"]; spot=data[name].get("spot"); r_fx=data[name].get("rate")
        spot_src=data[name].get("spot_src","—"); rate_src=data[name].get("rate_src","—")
        for ti,(tenor,days) in enumerate(TENORS):
            ws.row_dimensions[row].height=20
            is_last=(ti==n_tenor-1)
            bdr=Bdr()                                             # 主行始终细边
            bdr_last=Bdr("medium","1F3864") if is_last else Bdr() # 分行②行（币种块底边）
            fwd=calc_fwd(spot,r_usd,r_fx,days,ccy["day_basis"],ccy["type"])

            sc(ws,row,1,seq,font=F(8),align=A(),bg=bg,border=bdr,fmt="0")
            sc(ws,row,2,name if ti==0 else "",font=F(bold=True),align=A(),bg=bg,border=bdr)
            sc(ws,row,3,ccy["pair"] if ti==0 else "",font=F(8,color="404040"),align=A(),bg=bg,border=bdr)
            sc(ws,row,4,tenor,font=F(bold=True),align=A(),bg=bg,border=bdr)
            sc(ws,row,5,days,font=F(8),align=A(),bg=bg,border=bdr,fmt="0")

            if spot:
                fmt_s="0.0000" if name!="JPY" else "0.00"
                sc(ws,row,6,spot,font=F(9),align=A("right"),bg=C_INPUT,border=bdr,fmt=fmt_s)
            else:
                c=sc(ws,row,6,None,font=F(8,italic=True,color="E65100"),align=A(),bg=C_MISS,border=bdr)
                ws.cell(row,6).value="(询价)"

            if fwd:
                fmt_f="0.0000" if name!="JPY" else "0.00"
                sc(ws,row,7,fwd,font=F(9,color="1A5276"),align=A("right"),bg=C_CIP,border=bdr,fmt=fmt_f)
            else:
                sc(ws,row,7,None,font=F(8,italic=True,color="E65100"),align=A(),bg=C_MISS,border=bdr)
                ws.cell(row,7).value="(询价)"

            if ccy["type"]=="FX_BASE":
                hf=f'=IF(OR(ISNUMBER(F{row})=FALSE,ISNUMBER(G{row})=FALSE),"",(G{row}-F{row})/F{row}*360/E{row})'
            else:
                hf=f'=IF(OR(ISNUMBER(F{row})=FALSE,ISNUMBER(G{row})=FALSE),"",(F{row}-G{row})/F{row}*360/E{row})'
            sc(ws,row,8,hf,font=F(9),align=A(),bg=C_FORMULA,border=bdr,fmt="0.00%")

            dep=data[name].get("rate")
            if dep is not None:
                sc(ws,row,9,dep,font=F(9,bold=True,color="1B5E20"),align=A(),bg="F1F8E9",border=bdr,fmt="0.00%")
            else:
                sc(ws,row,9,None,font=F(8,italic=True,color="E65100"),align=A(),bg=C_MISS,border=bdr)

            if ccy["type"]=="FX_BASE":
                jf=(f'=IF(OR(ISNUMBER(F{row})=FALSE,ISNUMBER(G{row})=FALSE,'
                    f'ISNUMBER(I{row})=FALSE),"",((1+I{row}*E{row}/360)*G{row}/F{row}-1)*360/E{row})')
            else:
                jf=(f'=IF(OR(ISNUMBER(F{row})=FALSE,ISNUMBER(G{row})=FALSE,'
                    f'ISNUMBER(I{row})=FALSE),"",((1+I{row}*E{row}/360)*F{row}/G{row}-1)*360/E{row})')
            sc(ws,row,10,jf,font=F(9,bold=True,color=C_NAVY),align=A(),bg=C_FORMULA,border=bdr,fmt="0.00%")

            kf=f'=IF(J{row}="","",ROUND((J{row}-$F$3)*10000,1))'
            sc(ws,row,11,kf,font=F(9),align=A(),bg=C_FORMULA,border=bdr,fmt='+0.0" bp";-0.0" bp";"—"')

            sc(ws,row,12,f"汇率:{spot_src}\n利率:{rate_src}" if ti==0 else "",
               font=F(7,italic=True,color="595959"),align=A(wrap=True),bg=bg,border=bdr)

            if ti==0:
                if not spot: note="即期无数据，手动填"
                elif not fwd: note="远期需询价"
                elif dep is None: note="存款利率无数据，询价"
                else: note="远期=CIP理论；掉期用实际报价覆盖"
            else: note=""
            sc(ws,row,13,note,font=F(7,italic=True,color="595959"),align=A(wrap=True),bg=bg,border=bdr)
            seq+=1

            # ── 分行②示例行（同币种同期限，第二家境外分行询价槽）──
            b_row=row+1; prev=row
            bg_b="F3E5F5"
            fmt_sb="0.0000" if name!="JPY" else "0.00"
            ws.row_dimensions[b_row].height=16
            sc(ws,b_row,1,"↳",font=F(8,color="7B1FA2"),align=A(),bg=bg_b,border=bdr_last)
            sc(ws,b_row,2,"",font=F(8),align=A(),bg=bg_b,border=bdr_last)
            sc(ws,b_row,3,"",font=F(8),align=A(),bg=bg_b,border=bdr_last)
            sc(ws,b_row,4,tenor,font=F(7,italic=True,color="7B1FA2"),align=A(),bg=bg_b,border=bdr_last)
            sc(ws,b_row,5,days,font=F(8),align=A(),bg=bg_b,border=bdr_last,fmt="0")
            sc(ws,b_row,6,f"=F{prev}",font=F(9),align=A("right"),bg=C_INPUT,border=bdr_last,fmt=fmt_sb)
            sc(ws,b_row,7,f"=G{prev}",font=F(9,color="1A5276"),align=A("right"),bg=C_CIP,border=bdr_last,fmt=fmt_sb)
            sc(ws,b_row,8,f"=H{prev}",font=F(9),align=A(),bg=C_FORMULA,border=bdr_last,fmt="0.00%")
            sc(ws,b_row,9,None,font=F(8,italic=True,color="E65100"),align=A(),bg=C_MISS,border=bdr_last)
            if ccy["type"]=="FX_BASE":
                jf_b=(f'=IF(OR(ISNUMBER(F{b_row})=FALSE,ISNUMBER(G{b_row})=FALSE,'
                      f'ISNUMBER(I{b_row})=FALSE),"",((1+I{b_row}*E{b_row}/360)*G{b_row}/F{b_row}-1)*360/E{b_row})')
            else:
                jf_b=(f'=IF(OR(ISNUMBER(F{b_row})=FALSE,ISNUMBER(G{b_row})=FALSE,'
                      f'ISNUMBER(I{b_row})=FALSE),"",((1+I{b_row}*E{b_row}/360)*F{b_row}/G{b_row}-1)*360/E{b_row})')
            sc(ws,b_row,10,jf_b,font=F(9,bold=True,color=C_NAVY),align=A(),bg=C_FORMULA,border=bdr_last,fmt="0.00%")
            kf_b=f'=IF(J{b_row}="","",ROUND((J{b_row}-$F$3)*10000,1))'
            sc(ws,b_row,11,kf_b,font=F(9),align=A(),bg=C_FORMULA,border=bdr_last,fmt='+0.0" bp";-0.0" bp";"—"')
            sc(ws,b_row,12,"",font=F(8),align=A(),bg=bg_b,border=bdr_last)
            sc(ws,b_row,13,"↳ 分行②（示例，待询价）",font=F(7,italic=True,color="7B1FA2"),align=A(),bg=bg_b,border=bdr_last)

            row+=2

    n_all=len(CURRENCIES)*len(TENORS)*2
    ws.row_dimensions[row].height=6; row+=1
    ws.row_dimensions[row].height=22
    ws.merge_cells(f"A{row}:E{row}")
    sc(ws,row,1,"◆ 各期限最高综合USD收益",font=F(9,bold=True,color="FFFFFF"),align=A(),bg=C_NAVY,border=Bdr())
    for ti,(tenor,days) in enumerate(TENORS):
        tenor_rows=[]
        for ci in range(len(CURRENCIES)):
            r_main=DATA_START+ci*(n_tenor*2)+ti*2
            tenor_rows.extend([r_main, r_main+1])   # 主行+分行②
        j_refs=",".join(f"J{r}" for r in tenor_rows)
        sc(ws,row,6+ti,f'=IFERROR(MAX({j_refs}),"")',
           font=F(10,bold=True,color="C00000"),align=A(),bg="E8F5E9",border=Bdr(),fmt="0.00%")
    ws.merge_cells(f"J{row}:M{row}")
    sc(ws,row,10,"★ I列=公开基准参考，用实际询价覆盖后结果更准确",
       font=F(8,italic=True,color="E65100"),align=A("left",wrap=True),border=Bdr())

    k_range=f"K{DATA_START}:K{DATA_START+n_all-1}"
    ws.conditional_formatting.add(k_range,
        CellIsRule("greaterThan",["0"],
                   fill=PatternFill("solid",fgColor="C8E6C9"),
                   font=Font(name="微软雅黑",size=9,bold=True,color="1B5E20")))
    ws.conditional_formatting.add(k_range,
        CellIsRule("lessThan",["0"],
                   fill=PatternFill("solid",fgColor="FFCDD2"),
                   font=Font(name="微软雅黑",size=9,bold=True,color="B71C1C")))

    ws.freeze_panes="D7"
    ws.page_setup.orientation="landscape"
    ws.page_setup.fitToWidth=1
    ws.sheet_properties.tabColor=C_NAVY

    ws2=wb.create_sheet("数据来源",1)
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=3
    for col,w in zip("BCDEF",[16,14,14,12,28]):
        ws2.column_dimensions[col].width=w
    ws2.merge_cells("B1:F1")
    sc(ws2,1,2,f"实时抓取明细 — {today} {now}",font=F(12,bold=True,color=C_NAVY),align=A(),bg="EBF2FA")
    ws2.row_dimensions[1].height=26
    for j,h in enumerate(["类别","币种","抓取值","数据日期","状态","来源"],2):
        sc(ws2,2,j,h,font=F(9,bold=True,color="FFFFFF"),align=A(),bg=C_NAVY,border=Bdr())
    rows2=[]
    sofr2=data["SOFR"]
    rows2.append(("基准利率","SOFR",f"{sofr2['value']:.4%}" if sofr2["value"] else "ERR",
                  sofr2["date"]or"—","OK"if sofr2["value"]else"ERR","FRED/SOFR"))
    for ccy in CURRENCIES:
        nm=ccy["name"]; d2=data[nm]
        sp=d2.get("spot"); rt=d2.get("rate")
        dp=4 if nm!="JPY" else 2
        rows2.append(("即期",ccy["pair"],f"{sp:.{dp}f}"if sp else"ERR",d2.get("spot_date","—")or"—","OK"if sp else"ERR",d2.get("spot_src","—")))
        rows2.append(("利率",nm,f"{rt:.4%}"if rt else"估算",d2.get("rate_date","估算")or"估算","OK"if d2.get("rate_date","—")and d2.get("rate_date","—")!="估算"else"估算",ccy["rate_label"]))
    for ri,rd in enumerate(rows2,3):
        ws2.row_dimensions[ri].height=17
        bg2="F5F9FF" if ri%2==0 else "FFFFFF"
        for j2,v in enumerate(rd,2):
            c2=sc(ws2,ri,j2,v,font=F(9),align=A(),bg=bg2,border=Bdr())
            if j2==6:
                if "OK"in str(v): c2.fill=Fill("C8E6C9");c2.font=F(9,color="1B5E20")
                elif "ERR"in str(v): c2.fill=Fill("FFCDD2");c2.font=F(9,color="B71C1C")
                elif "估算"in str(v): c2.fill=Fill(C_MISS);c2.font=F(9,color="E65100")
    nr=len(rows2)+4
    ws2.merge_cells(f"B{nr}:F{nr}")
    sc(ws2,nr,2,"存款利率I列=公开基准参考，请向境外分行询价  远期G列=CIP理论值，请向掉期交易商询价",
       font=F(8,italic=True,color="E65100"),align=A("left"),bg="FFF3E0")
    ws2.sheet_properties.tabColor=C_BLUE

    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    return wb

def main():
    data=fetch_all_data()
    spot_ok=sum(1 for c in CURRENCIES if data[c["name"]].get("spot"))
    rate_ok=sum(1 for c in CURRENCIES if data[c["name"]].get("rate"))
    sofr_ok=data["SOFR"]["value"] is not None
    print(f"\n[汇总]  即期:{spot_ok}/{len(CURRENCIES)}  利率:{rate_ok}/{len(CURRENCIES)}  SOFR:{'OK'if sofr_ok else'ERR'}")
    print("\n生成Excel...")
    wb=build_excel(data)
    OUT_DIR.mkdir(parents=True,exist_ok=True)
    wb.save(str(OUT_FILE))
    print(f"OK：{OUT_FILE}")
    print("Sheet① 监控主表  Sheet② 数据来源")
    print("橙色I列=存款无数据 → 向境外分行询价手动填入")
    print("蓝色G列=CIP理论远期 → 向掉期交易商询价后覆盖")

if __name__=="__main__":
    main()