# -*- coding: utf-8 -*-
"""
merge_doha_rates.py
把「多哈报价」Sheet 的存款利率写入「监控主表」I列，M列加备注：
  I列 → ='多哈报价'!{列}{行}  （公式引用，链接到多哈报价Sheet）
  M列 → 备注"多哈分行"
  其他列（F/G/H/J/K等）一概不修改，保留原始公式/数值。
使用 openpyxl load_workbook 直接操作，保留原有格式。
"""

import pandas as pd
from openpyxl import load_workbook

FILE_PATH   = r'C:\Users\pynbj\OneDrive\1.积累要看的文件\1. 投资框架\11.投资机会跟踪报告\外币存款收益监控表_20260226.xlsx'
OUTPUT_PATH = r'C:\Users\pynbj\OneDrive\1.积累要看的文件\1. 投资框架\11.投资机会跟踪报告\外币存款收益监控表_20260226_updated.xlsx'

# ── 1. 读取多哈报价（pandas，用于 I 列数值写入）──────────────────────────────
doha_raw = pd.read_excel(FILE_PATH, sheet_name='多哈报价', header=0)
doha_raw.columns = [str(c).strip() for c in doha_raw.columns]
doha_raw.set_index('期限 (Tenor)', inplace=True)

print('=== 多哈报价 ===')
print(doha_raw.to_string())

DOHA_CURRENCIES = set(doha_raw.columns.tolist())
print(f'\n多哈报价币种：{DOHA_CURRENCIES}')

# ── 2. openpyxl 打开工作簿 ───────────────────────────────────────────────────
wb = load_workbook(FILE_PATH)
ws = wb['监控主表']

# 列索引（1-based）
COL_SEQ      = 1   # A  序号
COL_CURRENCY = 2   # B  币种
COL_PAIR     = 3   # C  报价对
COL_TENOR    = 4   # D  期限
COL_DAYS     = 5   # E  天数
COL_SPOT     = 6   # F  即期(Spot)
COL_FORWARD  = 7   # G  远期(Forward CIP理论)
COL_DEPOSIT  = 9   # I  存款报价(%) ★填询价
COL_NOTE     = 13  # M  分行/备注

# 数据起始行（第7行，跳过标题区）
DATA_START_ROW = 7

# 期限映射：监控主表 D列 → 多哈报价 index
TENOR_MAP = {'3M': '3 Mo', '6M': '6 Mo', '9M': '9 Mo', '12M': '12 Mo'}

# 多哈报价 Sheet 行号映射（期限 → 行号）
DOHA_ROW_MAP = {'3M': 7, '6M': 8, '9M': 9, '12M': 10}

# 多哈报价 Sheet 列字母映射（币种 → 列字母）
DOHA_COL_MAP = {
    'USD': 'B', 'CNH': 'C', 'EUR': 'D',
    'GBP': 'E', 'HKD': 'F', 'AUD': 'G', 'JPY': 'H',
}

# ── 3. G 列：写入 CIP 远期公式 ──────────────────────────────────────────────
# 多哈币种（EUR/GBP/AUD/JPY/HKD）：=F{r}*(1+'多哈报价'!B{dr}*E{r}/360)/(1+'多哈报价'!{fc}{dr}*E{r}/360)
# 非多哈币种（NZD/CAD/CHF/SGD）：=F{r}*(1+$F$3*E{r}/360)/(1+I{r}*E{r}/360)
print('\n── G列：写入 CIP 远期公式 ──')

fwd_count = 0
fwd_skip  = 0
current_currency = None

for row_num in range(DATA_START_ROW, ws.max_row + 1):
    seq_val      = ws.cell(row=row_num, column=COL_SEQ).value
    currency_val = ws.cell(row=row_num, column=COL_CURRENCY).value
    tenor_val    = ws.cell(row=row_num, column=COL_TENOR).value

    if currency_val and str(currency_val).strip() not in ('nan', ''):
        current_currency = str(currency_val).strip()

    if seq_val is None or str(seq_val).strip() in ('nan', '', '序'):
        continue

    tenor_str = str(tenor_val).strip() if tenor_val else ''
    doha_row  = DOHA_ROW_MAP.get(tenor_str)
    if not doha_row:
        continue

    r = row_num
    if current_currency and current_currency in DOHA_COL_MAP:
        # 多哈报价覆盖的币种：用多哈 USD 利率作基准
        fgn_col = DOHA_COL_MAP[current_currency]
        dr      = doha_row
        formula = (
            f"=F{r}*(1+'多哈报价'!B{dr}*E{r}/360)"
            f"/(1+'多哈报价'!{fgn_col}{dr}*E{r}/360)"
        )
        ws.cell(row=row_num, column=COL_FORWARD).value = formula
        fwd_count += 1
        print(f'  [OK] {r:3d}: {current_currency:4s} {tenor_str:4s}  {formula}')
    else:
        # 非多哈币种（NZD/CAD/CHF/SGD）：用 SOFR（$F$3）作 USD 基准
        formula = f'=F{r}*(1+$F$3*E{r}/360)/(1+I{r}*E{r}/360)'
        ws.cell(row=row_num, column=COL_FORWARD).value = formula
        fwd_skip += 1
        print(f'  ~    {r:3d}: {current_currency:4s} {tenor_str:4s}  {formula}  (SOFR基准)')

print(f'\n[OK] G列完成：写入 {fwd_count} 个多哈公式，{fwd_skip} 个SOFR公式')

# ── 4. I 列写入多哈存款报价公式 + M 列备注"多哈分行" ────────────────────────
print('\n── I列：写入多哈存款报价 / M列：备注 ──')

update_count = 0
skip_count   = 0
current_currency = None

for row_num in range(DATA_START_ROW, ws.max_row + 1):
    seq_val      = ws.cell(row=row_num, column=COL_SEQ).value
    currency_val = ws.cell(row=row_num, column=COL_CURRENCY).value
    tenor_val    = ws.cell(row=row_num, column=COL_TENOR).value

    if currency_val and str(currency_val).strip() not in ('nan', ''):
        current_currency = str(currency_val).strip()

    if seq_val is None or str(seq_val).strip() in ('nan', '', '序'):
        continue

    tenor_str  = str(tenor_val).strip() if tenor_val else ''
    doha_tenor = TENOR_MAP.get(tenor_str)
    if not doha_tenor:
        continue

    if not current_currency or current_currency not in DOHA_CURRENCIES:
        skip_count += 1
        print(f'  skip {row_num}: {current_currency} {tenor_str} — no data in doha')
        continue

    doha_row = DOHA_ROW_MAP.get(tenor_str)
    fgn_col  = DOHA_COL_MAP.get(current_currency)
    if not doha_row or not fgn_col:
        skip_count += 1
        continue

    # 写入公式：='多哈报价'!D7
    formula = f"='多哈报价'!{fgn_col}{doha_row}"
    ws.cell(row=row_num, column=COL_DEPOSIT).value = formula
    ws.cell(row=row_num, column=COL_NOTE).value    = '多哈分行'
    update_count += 1
    print(f'  [OK] {row_num:3d}: {current_currency:4s} {tenor_str:4s} -> I={formula} | 多哈分行')

print(f'\n? I列完成：更新 {update_count} 个，跳过 {skip_count} 个')

# ── 5. 保存 ─────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f'\n?? 已保存：{OUTPUT_PATH}')
