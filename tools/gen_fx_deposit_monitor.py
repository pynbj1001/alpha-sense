"""
外币存款综合收益监控表生成器
结构：客户USD → 境外掉期买外币 → 境外分行外币存款 → 到期掉期卖回USD

报价约定说明：
  EUR/USD, GBP/USD, AUD/USD, NZD/USD  → XXX/USD（外币是基础货币，价格=每1外币值多少USD）
    综合USD收益 = (Forward/Spot × (1+r×d/360) - 1) × 360/d
  USD/JPY, USD/HKD, USD/SGD, USD/CAD, USD/CHF → USD/XXX（USD是基础货币，价格=每1USD值多少外币）
    综合USD收益 = (Spot/Forward × (1+r×d/360) - 1) × 360/d

每周一更新，用于向客户推荐最优外币存款组合。
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from datetime import date
import os

# ─────────────────────────────────────────────
# 参数配置
# ─────────────────────────────────────────────
OUTPUT_DIR = r"c:\Users\pynbj\OneDrive\1.文档-积累要看的文件\1. 投资框架\11.投资机会跟踪报告"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"外币存款收益监控表_{date.today().strftime('%Y%m%d')}.xlsx")

TENORS = [
    ("3M",  90),
    ("6M",  180),
    ("9M",  270),
    ("12M", 360),
]

# 币种配置：(显示名, 报价对, 报价类型, 参考即期, 参考远期[3M,6M,9M,12M], 参考存款利率[3M,6M,9M,12M])
# 报价类型: "USD_BASE" = USD/XXX (如USD/JPY)；"FX_BASE" = XXX/USD (如EUR/USD)
# 参考数据仅作占位符，每周手动填入实盘报价
CURRENCIES = [
    {
        "name": "EUR",
        "pair": "EUR/USD",
        "type": "FX_BASE",
        "spot_ref": 1.0450,
        "fwd_ref":  [1.0490, 1.0530, 1.0570, 1.0610],   # 3M/6M/9M/12M 参考
        "dep_ref":  [3.20,   3.15,   3.10,   3.05],      # 欧元存款参考利率%
    },
    {
        "name": "GBP",
        "pair": "GBP/USD",
        "type": "FX_BASE",
        "spot_ref": 1.2650,
        "fwd_ref":  [1.2700, 1.2750, 1.2800, 1.2850],
        "dep_ref":  [4.50,   4.45,   4.40,   4.35],
    },
    {
        "name": "AUD",
        "pair": "AUD/USD",
        "type": "FX_BASE",
        "spot_ref": 0.6350,
        "fwd_ref":  [0.6315, 0.6280, 0.6245, 0.6210],
        "dep_ref":  [4.10,   4.05,   4.00,   3.95],
    },
    {
        "name": "NZD",
        "pair": "NZD/USD",
        "type": "FX_BASE",
        "spot_ref": 0.5650,
        "fwd_ref":  [0.5610, 0.5570, 0.5530, 0.5490],
        "dep_ref":  [4.30,   4.25,   4.20,   4.15],
    },
    {
        "name": "CAD",
        "pair": "USD/CAD",
        "type": "USD_BASE",
        "spot_ref": 1.4350,
        "fwd_ref":  [1.4300, 1.4250, 1.4200, 1.4150],
        "dep_ref":  [3.80,   3.75,   3.70,   3.65],
    },
    {
        "name": "CHF",
        "pair": "USD/CHF",
        "type": "USD_BASE",
        "spot_ref": 0.9050,
        "fwd_ref":  [0.9020, 0.8990, 0.8960, 0.8930],
        "dep_ref":  [0.80,   0.75,   0.70,   0.65],
    },
    {
        "name": "JPY",
        "pair": "USD/JPY",
        "type": "USD_BASE",
        "spot_ref": 149.50,
        "fwd_ref":  [148.00, 146.50, 145.00, 143.50],
        "dep_ref":  [0.10,   0.10,   0.12,   0.15],
    },
    {
        "name": "HKD",
        "pair": "USD/HKD",
        "type": "USD_BASE",
        "spot_ref": 7.7850,
        "fwd_ref":  [7.7840, 7.7830, 7.7820, 7.7810],
        "dep_ref":  [4.20,   4.15,   4.10,   4.05],
    },
    {
        "name": "SGD",
        "pair": "USD/SGD",
        "type": "USD_BASE",
        "spot_ref": 1.3450,
        "fwd_ref":  [1.3420, 1.3390, 1.3360, 1.3330],
        "dep_ref":  [3.50,   3.45,   3.40,   3.35],
    },
]

# ─────────────────────────────────────────────
# 样式配置
# ─────────────────────────────────────────────
COLOR_HEADER_DARK   = "1F3864"   # 深藏青
COLOR_HEADER_MID    = "2E5FA3"   # 中蓝
COLOR_HEADER_LIGHT  = "D6E4F0"   # 浅蓝
COLOR_CURRENCY_ODD  = "EBF2FA"   # 奇数币种浅蓝
COLOR_CURRENCY_EVEN = "FFFFFF"   # 偶数币种白
COLOR_INPUT         = "FFFDE7"   # 手工输入 - 淡黄
COLOR_FORMULA       = "F0FFF0"   # 公式计算 - 淡绿
COLOR_BEST          = "E8F5E9"   # 最优标记
COLOR_WARN          = "FFF3E0"   # 偏低提示

FONT_TITLE    = Font(name="微软雅黑", size=14, bold=True, color="1F3864")
FONT_HEADER   = Font(name="微软雅黑", size=9,  bold=True, color="FFFFFF")
FONT_SUBHDR   = Font(name="微软雅黑", size=8,  bold=True, color="1F3864")
FONT_NORMAL   = Font(name="微软雅黑", size=9)
FONT_BOLD     = Font(name="微软雅黑", size=9,  bold=True)
FONT_RED      = Font(name="微软雅黑", size=9,  color="C00000")
FONT_NOTE     = Font(name="Calibri", size=8, italic=True, color="595959")

ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT   = Alignment(horizontal="right",  vertical="center")

def thin_border(top=True, bottom=True, left=True, right=True):
    t = Side(style="thin", color="BFBFBF")
    n = Side(style=None)
    return Border(
        top=t if top else n,
        bottom=t if bottom else n,
        left=t if left else n,
        right=t if right else n,
    )

def thick_border_bottom():
    return Border(bottom=Side(style="medium", color="1F3864"))


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


# ─────────────────────────────────────────────
# 主函数
# ─────────────────────────────────────────────
def build_workbook():
    wb = openpyxl.Workbook()

    build_monitor_sheet(wb)
    build_history_sheet(wb)
    build_instruction_sheet(wb)

    # 删除默认Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"✅ 已生成：{OUTPUT_FILE}")


# ─────────────────────────────────────────────
# Sheet1：监控主表
# ─────────────────────────────────────────────
def build_monitor_sheet(wb):
    ws = wb.create_sheet("📊 监控主表", 0)
    ws.sheet_view.showGridLines = False

    # ── 列宽 ──
    col_widths = {
        "A": 6,   # 序号
        "B": 7,   # 币种
        "C": 7,   # 报价对
        "D": 6,   # 期限
        "E": 7,   # 天数
        # 以下分3个大组
        "F": 11,  # 即期汇率
        "G": 11,  # 远期汇率 (手填)
        "H": 10,  # 掉期年化%
        "I": 12,  # 境外存款利率%
        "J": 13,  # 综合USD收益%
        "K": 10,  # 较SOFR利差(bp)
        "L": 8,   # 存款来源
        "M": 22,  # 备注
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── 行高 ──
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 14
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 22

    # ─── 标题行 ───
    ws.merge_cells("A2:M2")
    title_cell = ws["A2"]
    title_cell.value = "外币存款综合收益监控表（境外掉期+境外分行存款）"
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_CENTER
    title_cell.fill = fill("EBF2FA")

    # ─── 更新信息行 ───
    ws["A3"] = "更新日期："
    ws["A3"].font = FONT_NOTE
    ws["B3"] = date.today().strftime("%Y-%m-%d")
    ws["B3"].font = Font(name="微软雅黑", size=9, bold=True, color="C00000")
    ws.merge_cells("B3:C3")

    ws["E3"] = "SOFR参考："
    ws["E3"].font = FONT_NOTE
    ws["F3"] = 4.30      # SOFR参考值（手动填）
    ws["F3"].number_format = "0.00%"
    ws["F3"].font = Font(name="微软雅黑", size=9, bold=True)
    ws["F3"].fill = fill(COLOR_INPUT)

    ws["G3"] = "（%，手动填写）"
    ws["G3"].font = FONT_NOTE
    ws.merge_cells("G3:H3")

    ws["J3"] = "制表："
    ws["J3"].font = FONT_NOTE
    ws["K3"] = "李XX / 交易部"
    ws["K3"].font = FONT_NOTE
    ws.merge_cells("K3:M3")

    # ─── 大组标题行 (row 4) ───
    group_headers = [
        ("A4:E4", "基本信息",     COLOR_HEADER_DARK),
        ("F4:H4", "掉期数据（境外）", COLOR_HEADER_MID),
        ("I4:I4", "存款数据",     COLOR_HEADER_MID),
        ("J4:K4", "综合收益",     COLOR_HEADER_DARK),
        ("L4:M4", "备注",         COLOR_HEADER_MID),
    ]
    for rng, txt, clr in group_headers:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = txt
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.fill = fill(clr)

    # ─── 列子标题行 (row 5) ───
    sub_headers = [
        ("A5", "序号"),
        ("B5", "币种"),
        ("C5", "报价对"),
        ("D5", "期限"),
        ("E5", "天数"),
        ("F5", "即期汇率\n(Spot)"),
        ("G5", "远期汇率\n(Forward)"),
        ("H5", "掉期年化\n(%)"),
        ("I5", "境外存款\n报价(%)"),
        ("J5", "综合USD\n收益(%)"),
        ("K5", "较SOFR\n利差(bp)"),
        ("L5", "报价来源"),
        ("M5", "备注"),
    ]
    for addr, txt in sub_headers:
        c = ws[addr]
        c.value = txt
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.fill = fill(COLOR_HEADER_DARK)
        c.border = thin_border()

    # ─── 说明行 (row 6) ───
    ws.merge_cells("A6:E6")
    ws["A6"].value = "↓ 黄色=手动填入  绿色=自动计算"
    ws["A6"].font = FONT_NOTE
    ws["A6"].alignment = ALIGN_LEFT

    ws.merge_cells("F6:G6")
    ws["F6"].value = "即期/远期：填市场报价（同口径）"
    ws["F6"].font = FONT_NOTE
    ws["F6"].alignment = ALIGN_CENTER

    ws["H6"].value = "=(远期-即期)/即期×360/天数 或 -(远期-即期)/即期×360/天数（见说明表）"
    ws["H6"].font = FONT_NOTE
    ws.merge_cells("H6:J6")

    ws.merge_cells("K6:M6")
    ws["K6"].value = "综合USD收益 - SOFR（单位：基点）"
    ws["K6"].font = FONT_NOTE

    # ─── 数据行 ───
    DATA_START_ROW = 7
    row = DATA_START_ROW
    seq = 1

    for ci, ccy in enumerate(CURRENCIES):
        bg_color = COLOR_CURRENCY_ODD if ci % 2 == 0 else COLOR_CURRENCY_EVEN

        for ti, (tenor, days) in enumerate(TENORS):
            first_in_group = (ti == 0)
            last_in_group  = (ti == len(TENORS) - 1)

            ws.row_dimensions[row].height = 18

            # 序号
            ws[f"A{row}"] = seq
            ws[f"A{row}"].font = FONT_NORMAL
            ws[f"A{row}"].alignment = ALIGN_CENTER
            ws[f"A{row}"].fill = fill(bg_color)

            # 币种（仅首行显示，合并）
            if first_in_group:
                ws[f"B{row}"] = ccy["name"]
                ws[f"B{row}"].font = FONT_BOLD
                ws[f"B{row}"].alignment = ALIGN_CENTER
                ws[f"B{row}"].fill = fill(bg_color)
                # 报价对
                ws[f"C{row}"] = ccy["pair"]
                ws[f"C{row}"].font = Font(name="Calibri", size=8, color="404040")
                ws[f"C{row}"].alignment = ALIGN_CENTER
                ws[f"C{row}"].fill = fill(bg_color)
            else:
                ws[f"B{row}"].fill = fill(bg_color)
                ws[f"B{row}"].font = FONT_NORMAL
                ws[f"C{row}"].fill = fill(bg_color)
                ws[f"C{row}"].font = FONT_NORMAL

            # 期限 / 天数
            ws[f"D{row}"] = tenor
            ws[f"D{row}"].font = FONT_BOLD
            ws[f"D{row}"].alignment = ALIGN_CENTER
            ws[f"D{row}"].fill = fill(bg_color)

            ws[f"E{row}"] = days
            ws[f"E{row}"].font = FONT_NORMAL
            ws[f"E{row}"].alignment = ALIGN_CENTER
            ws[f"E{row}"].fill = fill(bg_color)

            # 即期汇率 (手填，预置参考值)
            ws[f"F{row}"] = ccy["spot_ref"]
            ws[f"F{row}"].number_format = (
                "0.0000" if ccy["type"] == "FX_BASE" else
                ("0.00" if ccy["name"] == "JPY" else "0.0000")
            )
            ws[f"F{row}"].alignment = ALIGN_RIGHT
            ws[f"F{row}"].fill = fill(COLOR_INPUT)
            ws[f"F{row}"].font = FONT_NORMAL

            # 远期汇率 (手填)
            ws[f"G{row}"] = ccy["fwd_ref"][ti]
            ws[f"G{row}"].number_format = ws[f"F{row}"].number_format
            ws[f"G{row}"].alignment = ALIGN_RIGHT
            ws[f"G{row}"].fill = fill(COLOR_INPUT)
            ws[f"G{row}"].font = FONT_NORMAL

            # 掉期年化% (公式)
            # FX_BASE (EUR/USD etc.): swap_ann = (G-F)/F × 360/E
            # USD_BASE (USD/JPY etc.): swap_ann = (F-G)/F × 360/E  (远期USD/JPY跌=USD走弱=获益)
            if ccy["type"] == "FX_BASE":
                h_formula = f"=(G{row}-F{row})/F{row}*360/E{row}"
            else:
                h_formula = f"=(F{row}-G{row})/F{row}*360/E{row}"
            ws[f"H{row}"] = h_formula
            ws[f"H{row}"].number_format = "0.00%"
            ws[f"H{row}"].alignment = ALIGN_CENTER
            ws[f"H{row}"].fill = fill(COLOR_FORMULA)
            ws[f"H{row}"].font = FONT_NORMAL

            # 境外存款利率% (手填)
            ws[f"I{row}"] = ccy["dep_ref"][ti] / 100   # 存为小数，显示为%
            ws[f"I{row}"].number_format = "0.00%"
            ws[f"I{row}"].alignment = ALIGN_CENTER
            ws[f"I{row}"].fill = fill(COLOR_INPUT)
            ws[f"I{row}"].font = FONT_NORMAL

            # 综合USD收益% (公式，年化简单利率)
            # FX_BASE:  ((1+I×E/360)×G/F - 1) × 360/E
            # USD_BASE: ((1+I×E/360)×F/G - 1) × 360/E
            if ccy["type"] == "FX_BASE":
                j_formula = f"=((1+I{row}*E{row}/360)*G{row}/F{row}-1)*360/E{row}"
            else:
                j_formula = f"=((1+I{row}*E{row}/360)*F{row}/G{row}-1)*360/E{row}"
            ws[f"J{row}"] = j_formula
            ws[f"J{row}"].number_format = "0.00%"
            ws[f"J{row}"].alignment = ALIGN_CENTER
            ws[f"J{row}"].fill = fill(COLOR_FORMULA)
            ws[f"J{row}"].font = Font(name="微软雅黑", size=9, bold=True)

            # 较SOFR利差 (bp)
            # SOFR 在 F3，需转为小数
            ws[f"K{row}"] = f"=(J{row}-F3)*10000"
            ws[f"K{row}"].number_format = '0" bp"'
            ws[f"K{row}"].alignment = ALIGN_CENTER
            ws[f"K{row}"].fill = fill(COLOR_FORMULA)
            ws[f"K{row}"].font = FONT_NORMAL

            # 来源 (手填)
            ws[f"L{row}"] = "境外分行"
            ws[f"L{row}"].font = FONT_NOTE
            ws[f"L{row}"].alignment = ALIGN_CENTER
            ws[f"L{row}"].fill = fill(bg_color)

            # 备注 (手填)
            ws[f"M{row}"].fill = fill(bg_color)

            # ── 边框 ──
            for col in "ABCDEFGHIJKLM":
                cell = ws[f"{col}{row}"]
                if last_in_group:
                    cell.border = Border(
                        top=Side(style="thin", color="BFBFBF"),
                        bottom=Side(style="medium", color="1F3864"),
                        left=Side(style="thin", color="BFBFBF"),
                        right=Side(style="thin", color="BFBFBF"),
                    )
                else:
                    cell.border = thin_border()

            seq += 1
            row += 1

    # ── 汇总行：各期限最优收益参考 ──
    row += 1
    ws.row_dimensions[row].height = 8
    row += 1
    summary_row = row
    ws.merge_cells(f"A{summary_row}:E{summary_row}")
    ws[f"A{summary_row}"] = "各期限最高综合USD收益（自动）"
    ws[f"A{summary_row}"].font = Font(name="微软雅黑", size=9, bold=True, color="1F3864")
    ws[f"A{summary_row}"].alignment = ALIGN_CENTER
    ws[f"A{summary_row}"].fill = fill(COLOR_HEADER_LIGHT)
    ws[f"A{summary_row}"].border = thin_border()

    j_col_data = [f"J{DATA_START_ROW + ci*4 + ti}" for ci in range(len(CURRENCIES)) for ti in range(4)]
    for ti, (tenor, days) in enumerate(TENORS):
        # collect all J values for this tenor
        j_refs = [f"J{DATA_START_ROW + ci*4 + ti}" for ci in range(len(CURRENCIES))]
        j_list = ",".join(j_refs)
        col_offset = ti  # F onwards for each tenor
        col_letter = get_column_letter(6 + ti)   # F=3M, G=6M, H=9M, I=12M
        ws.row_dimensions[summary_row].height = 22
        target_cell = ws[f"{col_letter}{summary_row}"]
        target_cell.value = f"=MAX({j_list})"
        target_cell.number_format = "0.00%"
        target_cell.font = Font(name="微软雅黑", size=9, bold=True, color="C00000")
        target_cell.alignment = ALIGN_CENTER
        target_cell.fill = fill(COLOR_BEST)
        target_cell.border = thin_border()

    # 标注期限标签
    for ti, (tenor, _) in enumerate(TENORS):
        col_letter = get_column_letter(6 + ti)
        ws.row_dimensions[summary_row - 1].height = 16
        label_cell = ws[f"{col_letter}{summary_row - 1}"]
        label_cell.value = tenor
        label_cell.font = FONT_SUBHDR
        label_cell.alignment = ALIGN_CENTER
        label_cell.fill = fill(COLOR_HEADER_LIGHT)
        label_cell.border = thin_border()

    # ── 条件格式：综合USD收益 > SOFR 时高亮 ──
    j_range = f"J{DATA_START_ROW}:J{DATA_START_ROW + len(CURRENCIES)*4 - 1}"
    ws.conditional_formatting.add(
        j_range,
        ColorScaleRule(
            start_type="min", start_color="FF4444",
            mid_type="percentile", mid_value=50, mid_color="FFFF00",
            end_type="max", end_color="00AA00"
        )
    )

    # ── 冻结窗口 ──
    ws.freeze_panes = "D7"

    # ── 打印设置 ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "4:5"


# ─────────────────────────────────────────────
# Sheet2：历史记录
# ─────────────────────────────────────────────
def build_history_sheet(wb):
    ws = wb.create_sheet("📅 历史记录", 1)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 6
    for col in "DEFGHIJK":
        ws.column_dimensions[col].width = 12

    # 标题
    ws.merge_cells("A1:K1")
    ws["A1"].value = "外币存款收益历史追踪"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_CENTER
    ws["A1"].fill = fill("EBF2FA")
    ws.row_dimensions[1].height = 28

    headers = ["日期", "币种", "期限", "即期汇率", "远期汇率", "掉期年化%",
               "存款利率%", "综合USD收益%", "SOFR%", "利差(bp)", "备注"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.fill = fill(COLOR_HEADER_DARK)
        c.border = thin_border()
    ws.row_dimensions[2].height = 22

    # 示例数据占位
    ws["A3"] = "← 每周将当周数据粘贴至此处存档"
    ws["A3"].font = Font(name="微软雅黑", size=9, italic=True, color="999999")
    ws.merge_cells("A3:K3")
    ws.row_dimensions[3].height = 18

    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────
# Sheet3：说明
# ─────────────────────────────────────────────
def build_instruction_sheet(wb):
    ws = wb.create_sheet("📋 使用说明", 2)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    ws.merge_cells("A1:B1")
    ws["A1"].value = "外币存款收益监控表 — 使用说明"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_CENTER
    ws["A1"].fill = fill("EBF2FA")
    ws.row_dimensions[1].height = 28

    instructions = [
        ("交易结构", "客户USD资金 → ①境外掉期：卖USD买外币（近端） → ②存入境外分行外币存款 → ③到期掉期：卖外币买USD（远端），全程汇率风险锁定"),
        ("掉期来源", "境外机构报价（境内先走北京分行，但金市暂无授信，优先使用境外掉期）"),
        ("存款品种", "优先选择外币活期/定期存款（Deposit）；CD流动性较低，需凑单，持有至到期"),
        ("报价约定-FX_BASE", "EUR/USD、GBP/USD、AUD/USD、NZD/USD\n报价含义：1外币=X USD\n掉期年化=(远期-即期)/即期×360/天数\n综合USD收益=((1+存款利率×天数/360)×远期/即期-1)×360/天数"),
        ("报价约定-USD_BASE", "USD/JPY、USD/HKD、USD/SGD、USD/CAD、USD/CHF\n报价含义：1 USD=X外币\n掉期年化=(即期-远期)/即期×360/天数\n综合USD收益=((1+存款利率×天数/360)×即期/远期-1)×360/天数"),
        ("手动填写项（黄色）", "F列：即期汇率（Spot，每日盘中实时报价）\nG列：远期汇率（Forward Outright，与Spot同口径）\nI列：境外分行外币存款利率（%，询价后填入）\nF3：当周SOFR参考利率（%）"),
        ("自动计算项（绿色）", "H列：掉期年化收益/成本（%）\nJ列：综合USD收益率（%，年化简单利率）\nK列：较SOFR利差（基点，bp）"),
        ("更新频率", "每周一开盘后更新，将上周数据复制至【历史记录】存档"),
        ("注意事项", "①远期汇率务必与即期汇率使用相同报价口径\n②存款利率填写年化利率（%），如 4.50 表示年化4.50%\n③综合收益为年化估算，实际以交割金额为准\n④CD品种需注意流动性溢价，通常高于普通存款20-50bp"),
    ]

    for ri, (key, val) in enumerate(instructions, 2):
        ws.row_dimensions[ri + 1].height = max(30, val.count("\n") * 14 + 16)
        kc = ws.cell(row=ri + 1, column=1, value=key)
        kc.font = FONT_BOLD
        kc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        kc.fill = fill(COLOR_HEADER_LIGHT)
        kc.border = thin_border()

        vc = ws.cell(row=ri + 1, column=2, value=val)
        vc.font = FONT_NORMAL
        vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        vc.border = thin_border()


# ─────────────────────────────────────────────
if __name__ == "__main__":
    build_workbook()
